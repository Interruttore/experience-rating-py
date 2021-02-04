from zipfile import error
from openpyxl import load_workbook
from datetime import date
import constants
import tools


def addMovie(info, filename):

    try:
        workbook = load_workbook(filename=filename)
        sheet = workbook[constants.MOVIE_SHEET]

    except:
        # TODO close workbook
        print("Error in the excel")
        return error

    today = date.today()
    d1 = today.strftime("%d/%m/%Y")
    index = tools.getIndex(sheet)

    try:
        sheet.cell(index, constants.MOVIE_NAME).value = info["name"]
        sheet.cell(
            index, constants.MOVIE_RELEASE_DATE).value = info["release_date"]
        sheet.cell(index, constants.MOVIE_REVIEW_DATE).value = d1
        sheet.cell(index, constants.MOVIE_GENRE).value = info["genre"]
        sheet.cell(index, constants.MOVIE_VOTE).value = info["vote"]
    except:
        print("Error in adding the movie")
    workbook.acive = workbook.save(filename=filename)

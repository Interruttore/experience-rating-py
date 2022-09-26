from openpyxl import load_workbook
from datetime import date
import constants


# def get_index(sheet):
#     i = 0
#     index = 0
#     for value in sheet.iter_rows(min_row=3, min_col=1, max_col=1, values_only=True):
#         if(value[0] != None):
#             i += 1
#         else:
#             index = i+3
#             break
#     return index


def add_movie(info, filename):

    print("ADD_MOVIE:")

    try:
        workbook = load_workbook(filename=filename)
        sheet = workbook[constants.MOVIE_SHEET]

    except:
        # TODO close workbook
        raise Exception(constants.EXCEL_OPEN_ERR)

    today = date.today()
    d1 = today.strftime("%d/%m/%Y")
    #index = get_index(sheet)

    try:

        sheet.append([info["name"], info["release_date"],
                      d1, info["genre"], float(info["vote"])])
        # sheet.cell(index, constants.MOVIE_NAME).value = info["name"]
        # sheet.cell(
        #     index, constants.MOVIE_RELEASE_DATE).value = info["release_date"]
        # sheet.cell(index, constants.MOVIE_REVIEW_DATE).value = d1
        # sheet.cell(index, constants.MOVIE_GENRE).value = info["genre"]
        # sheet.cell(index, constants.MOVIE_VOTE).value = float(info["vote"])
    except:
        print("Error in adding the movie")
    workbook.acive = workbook.save(filename=filename)


def remove_duplicate(filename):

    try:
        workbook = load_workbook(filename=filename)
        sheet = workbook[constants.MOVIE_SHEET]

    except Exception as e:
        print("Error:", e)

    try:
        for row in sheet.iter_rows(min_row=1):
            print(row[0].value)
    except Exception as e:
        print("Error:", e)
